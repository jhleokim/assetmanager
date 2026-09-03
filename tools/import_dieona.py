#!/usr/bin/env python3
"""디어나 가계부 ver7.x (.xlsx) → 이 앱의 JSON 백업 형식.

사용:  python3 tools/import_dieona.py 가계부.xlsx > ledger.json
       python3 tools/import_dieona.py 가계부.xlsx --sample   # '월별시트 샘플페이지'도 읽는다 (검증용)
결과 JSON은 앱 [설정] → [JSON 불러오기]로 병합한다 (같은 id는 최신 쪽, 나머지는 합집합).

옮기는 것
  설정        지출방법(C2:L2) · 시작 년/월/일(C3,E3,G3) · 대분류/소분류(B6:Q29)
  월별시트    이달의 고정지출 내역(T7:Z26) · 이달의 수입/저축/지출 내역(T30:Z~) · 태그(E15:E24)
  예비비      입금 내역(B26:D~) → 저축/예비비, 지출 내역(K10:O~) → 예비비/분류, 분류별 예산(B10:C21)
  결제일 관리  → 고정지출 규칙 (금액·결제일이 있는 행만)
검증 규칙은 템플릿과 같다: 날짜·금액·대분류·소분류 중 하나라도 비면 그 행은 버린다 (stderr에 개수 표시).
'자산관리'·'통장관리'·'카드 관리' 시트는 옮기지 않는다 — 자산은 이 앱의 [자산 목록]이 담당한다."""
import sys, json, re, datetime as dt
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 이 필요합니다:  pip install openpyxl")

FUND_CAT, FUND_SUB = "예비비", "예비비"
KIND_OF = {"수입": "income", "저축": "saving"}

def iso(v):
    if isinstance(v, (dt.datetime, dt.date)): return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)) and v > 20000:           # 엑셀 일련번호
        return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(v))).strftime("%Y-%m-%d")
    if isinstance(v, str):
        m = re.match(r"^\s*(\d{4})[.\-/년 ]+(\d{1,2})[.\-/월 ]+(\d{1,2})", v)
        if m: return "%04d-%02d-%02d" % tuple(int(x) for x in m.groups())
    return None

def num(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return int(round(v))
    t = re.sub(r"[,\s원]", "", str(v))
    try: return int(round(float(t)))
    except ValueError: return None

def txt(v): return "" if v is None else str(v).strip()

def read_settings(ws):
    methods = [txt(ws.cell(2, c).value) for c in range(3, 13)]
    methods = [m for m in methods if m]
    y, m, d = num(ws["C3"].value) or dt.date.today().year, num(ws["E3"].value) or 1, num(ws["G3"].value) or 1
    cats = []
    for r in range(6, 30):
        name = txt(ws.cell(r, 2).value)
        if not name: continue
        subs = [txt(ws.cell(r, c).value) for c in range(3, 18)]
        subs = [s for s in subs if s]
        cats.append({"name": name, "kind": KIND_OF.get(name, "expense"), "subs": subs, **({"fixed": True} if name == "고정지출" else {})})
    return methods, (y, m, d), cats

def read_entries(ws, cats_by_name, log):
    """T:날짜 U:내용 V:금액 W:대분류 X:소분류 Y:지출방법 Z:태그. 두 표(7~26, 30~) 모두."""
    out, skipped = [], 0
    for r in list(range(7, 27)) + list(range(30, ws.max_row + 1)):
        date, memo, amt, cat, sub, met, tag = (ws.cell(r, c).value for c in range(20, 27))
        if all(v in (None, "") for v in (date, memo, amt, sub)) : continue        # 빈 줄 (W열 '고정지출' 선입력은 무시)
        d, a, c, s = iso(date), num(amt), txt(cat), txt(sub)
        if not (d and a and c and s) or c not in cats_by_name or s not in cats_by_name[c]["subs"]:
            skipped += 1; continue
        cinfo = cats_by_name[c]
        out.append({"date": d, "amount": a, "kind": cinfo["kind"], "cat": c, "sub": s, "memo": txt(memo), "method": txt(met), "tag": txt(tag), "fund": False})
    if skipped: log.append("%s: 필수 항목이 빈 %d행 제외" % (ws.title, skipped))
    return out

def read_tags(ws):
    return [txt(ws.cell(r, 5).value) for r in range(15, 25) if txt(ws.cell(r, 5).value)]

def read_fund(ws, log):
    entries, budgets, subs = [], {}, []
    for r in range(10, 22):
        name = txt(ws.cell(r, 2).value)
        if not name: continue
        subs.append(name); b = num(ws.cell(r, 3).value)
        if b: budgets[name] = b
    skipped = 0
    for r in range(26, ws.max_row + 1):                        # 입금: B날짜 C금액 D내용
        d, a = iso(ws.cell(r, 2).value), num(ws.cell(r, 3).value)
        if d and a: entries.append({"date": d, "amount": a, "kind": "saving", "cat": "저축", "sub": FUND_SUB, "memo": txt(ws.cell(r, 4).value), "method": "", "tag": "", "fund": False})
        elif d or a: skipped += 1
    for r in range(10, ws.max_row + 1):                        # 지출: K날짜 L내용 M금액 N분류 O비고
        d, memo, a, cat, note = (ws.cell(r, c).value for c in range(11, 16))
        d, a, c = iso(d), num(a), txt(cat)
        if d and a and c:
            if c not in subs: subs.append(c)
            entries.append({"date": d, "amount": a, "kind": "expense", "cat": FUND_CAT, "sub": c, "memo": (txt(memo) + (" · " + txt(note) if txt(note) else "")).strip(), "method": "", "tag": "", "fund": True})
        elif d or a or c: skipped += 1
    if skipped: log.append("예비비: 필수 항목이 빈 %d행 제외" % skipped)
    return entries, budgets, subs

def read_recurring(ws, cats_by_name, log):
    rules, skipped = [], 0
    for r in range(4, ws.max_row + 1):
        name, kind, amt, day, met, note = (ws.cell(r, c).value for c in range(3, 9))
        if not txt(name) and not amt: continue
        a = num(amt); m = re.search(r"(\d{1,2})", txt(day))
        if not a or not m: skipped += 1; continue
        cat = "고정지출" if txt(kind) in ("", "고정지출", "정기결제") else txt(kind)
        cinfo = cats_by_name.get(cat) or cats_by_name.get("고정지출")
        if not cinfo: skipped += 1; continue
        sub = guess_sub(txt(name), cinfo["subs"])
        rules.append({"name": txt(name), "amount": a, "cat": cinfo["name"], "sub": sub, "day": int(m.group(1)), "method": txt(met), "memo": txt(note), "active": True})
    if skipped: log.append("결제일 관리: 금액·결제일이 없는 %d행 제외" % skipped)
    return rules

def guess_sub(name, subs):
    hints = {"보험": "보험료", "통신": "통신비", "폰": "통신비", "인터넷": "통신비", "월세": "주거비", "관리비": "주거비", "대출": "주거비", "교통": "교통비", "주유": "교통비"}
    for k, v in hints.items():
        if k in name and v in subs: return v
    return subs[0]

def main():
    if len(sys.argv) < 2: sys.exit(__doc__)
    path, want_sample = sys.argv[1], "--sample" in sys.argv
    wb = openpyxl.load_workbook(path, data_only=True)
    log = []
    methods, (y, m, d), cats = read_settings(wb["설정"])
    if not any(c["name"] == FUND_CAT for c in cats):
        cats.append({"name": FUND_CAT, "kind": "expense", "fund": True, "subs": []})
    sv = next((c for c in cats if c["kind"] == "saving"), None)
    if sv and FUND_SUB not in sv["subs"]: sv["subs"].append(FUND_SUB)
    by = {c["name"]: c for c in cats}
    entries, tags = [], []
    sheets = [str(i) for i in range(1, 13)] + (["월별시트 샘플페이지"] if want_sample else [])
    for nm in sheets:
        if nm in wb.sheetnames:
            entries += read_entries(wb[nm], by, log)
            if not tags: tags = read_tags(wb[nm])
    fund_entries, fund_budgets, fund_subs = ([], {}, [])
    if "예비비" in wb.sheetnames:
        fund_entries, fund_budgets, fund_subs = read_fund(wb["예비비"], log)
        by[FUND_CAT]["subs"] = list(dict.fromkeys(by[FUND_CAT]["subs"] + fund_subs))
    rules = read_recurring(wb["결제일 관리"], by, log) if "결제일 관리" in wb.sheetnames else []
    now = int(dt.datetime.now().timestamp() * 1000)
    seq = 10_000_000                                            # 앱 id 와 겹치지 않게 큰 수부터 — 병합 시 seq 가 이어진다
    def stamp(x):
        nonlocal seq; seq += 1; x["id"] = seq; x["updatedAt"] = now; return x
    out = {"v": 3, "seq": seq + 1, "assets": [], "trades": [], "hist": {}, "snaps": [], "set": {}, "tomb": {},
           "ledger": {"settings": {"startDay": max(1, min(28, d)), "methods": methods, "tags": tags or ["소비", "투자", "반성"], "cats": cats,
                                   "budgets": {}, "fundBudgets": fund_budgets, "updatedAt": now},
                      "entries": [stamp(e) for e in entries + fund_entries], "recurring": [stamp(r) for r in rules], "events": []}}
    out["seq"] = seq + 1
    tot = lambda k: sum(e["amount"] for e in out["ledger"]["entries"] if e["kind"] == k and not e["fund"])
    log.append("가져옴: 대분류 %d · 항목 %d (수입 %s / 지출 %s / 저축 %s / 예비비 지출 %s) · 규칙 %d · 시작일 %d일 (%d년 %d월부터)" % (
        len(cats), len(out["ledger"]["entries"]), f"{tot('income'):,}", f"{tot('expense'):,}", f"{tot('saving'):,}",
        f"{sum(e['amount'] for e in out['ledger']['entries'] if e['fund']):,}", len(rules), d, y, m))
    print("\n".join(log), file=sys.stderr)
    json.dump(out, sys.stdout, ensure_ascii=False, indent=1)

if __name__ == "__main__":
    main()
